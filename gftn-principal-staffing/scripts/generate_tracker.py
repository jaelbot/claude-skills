"""
GFTN Principal Staffing — Schedule Tracker Generator
Usage: python generate_tracker.py --config event_config.json --output tracker.xlsx

Palette and structure aligned to GFTN brand (navy) and the Principal_Schedule Tracker.xlsx template.
"""
import argparse, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── GFTN Palette ──────────────────────────────────────────────────────────────
NAVY_DARK   = "032C68"   # main header fill (GFTN brand navy)
NAVY_MID    = "0F4761"   # section headers
NAVY_LIGHT  = "D6E4F0"   # light navy tint for alternating rows
WHITE       = "FFFFFF"
PALE_GREY   = "F5F5F5"   # alternating row background
LIGHT_GREY  = "E8E8E8"
DARK_GREY   = "3A3A3A"
AMBER       = "FFC107"   # warning / note tint

YELLOW      = "FFF9C4"   # Suggested activities
ORANGE      = "FFE0B2"   # Transfer / Meal rows

# Per-principal colour columns in Master Overview (matches template exactly)
PRINCIPAL_COLOURS = [
    "D0E8DB",   # light green  — POC 1 / Coordinator
    "E3F2FD",   # light blue   — POC 2
    "FFF9C4",   # light yellow — POC 3
    "FCE4EC",   # light pink   — POC 4
    "EDE7F6",   # light purple — POC 5
    "FBE9E7",   # light coral  — POC 6
]

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _header_row(ws, row, headers, widths, fill_color=NAVY_MID, height=28):
    ws.row_dimensions[row].height = height
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, size=10, color=WHITE)
        c.fill = _fill(fill_color)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w


def _title_row(ws, text, span, fill_color=NAVY_DARK, height=32, font_size=13):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]
    c.value = text
    c.font = _font(bold=True, size=font_size, color=WHITE)
    c.fill = _fill(fill_color)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = height


def _add_dv(ws, formula, cells):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    for cell in cells:
        dv.add(cell)


def _row_fill(idx, type_val=""):
    if type_val == "Suggested":
        return YELLOW
    if type_val in ("Transfer", "Meal"):
        return ORANGE
    return NAVY_LIGHT if idx % 2 == 0 else WHITE


# ── MASTER OVERVIEW ────────────────────────────────────────────────────────────

def build_master_overview(wb, cfg):
    ws = wb.active
    ws.title = "Master Overview"
    ws.freeze_panes = "B4"

    principals = cfg["principals"]
    num_principals = len(principals)
    total_cols = 5 + num_principals  # Date, Start, End, Activity, Type + one col per principal

    _title_row(ws,
        f"{cfg['event_name']} — Principal Schedule Overview  |  {cfg['dates']}  |  {cfg['location']}",
        total_cols)

    # Instruction row
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    c = ws["A2"]
    c.value = (
        f"Managed by: {cfg.get('coordinator','Main Coordinator')}   |   "
        "Colour-coded by principal / POC assignment   |   "
        "Upload to SharePoint and share with Edit access for live co-authoring   |   "
        "Keep updated throughout the event"
    )
    c.font = _font(italic=True, size=9, color=DARK_GREY)
    c.fill = _fill(NAVY_LIGHT)
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 22

    # Column headers — base columns use navy, principal columns use their assigned colour
    base_headers = ["Date", "Time (Start)", "Time (End)", "Activity / Session", "Type"]
    base_widths  = [14, 13, 13, 38, 18]
    p_headers    = [f"{p['name']}\n(POC: {p['poc']})" for p in principals]
    p_widths     = [22] * num_principals

    ws.row_dimensions[3].height = 38
    for col, (h, w) in enumerate(zip(base_headers + p_headers, base_widths + p_widths), start=1):
        c = ws.cell(row=3, column=col, value=h)
        if col <= 5:
            c.font = _font(bold=True, size=10, color=WHITE)
            c.fill = _fill(NAVY_MID)
        else:
            p_idx = col - 6
            pc = PRINCIPAL_COLOURS[p_idx % len(PRINCIPAL_COLOURS)]
            c.font = _font(bold=True, size=10, color=NAVY_DARK)
            c.fill = _fill(pc)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    # Aggregate all timeslots
    slots = {}
    for p in principals:
        for eng in p.get("engagements", []):
            key = (eng.get("date",""), eng.get("time_start",""), eng.get("time_end",""),
                   eng.get("activity",""), eng.get("type",""))
            if key not in slots:
                slots[key] = {}
            slots[key][p["name"]] = True

    for r_idx, (key, p_map) in enumerate(sorted(slots.items()), start=4):
        date, ts, te, act, typ = key
        row_bg = _row_fill(r_idx, typ)
        for col, val in enumerate([date, ts, te, act, typ], start=1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.font = _font(size=10, color=DARK_GREY)
            c.fill = _fill(row_bg)
            c.alignment = _align("center" if col <= 3 else "left")
            c.border = _border()
        for p_idx, p in enumerate(principals):
            col = 6 + p_idx
            attending = "\u2713" if p["name"] in p_map else ""
            c = ws.cell(row=r_idx, column=col, value=attending)
            c.font = _font(bold=bool(attending), size=12, color=NAVY_DARK if attending else LIGHT_GREY)
            pc = PRINCIPAL_COLOURS[p_idx % len(PRINCIPAL_COLOURS)]
            c.fill = _fill(pc if attending else WHITE)
            c.alignment = _align("center")
            c.border = _border()
        ws.row_dimensions[r_idx].height = 22

    # Blank template rows when no engagements yet
    if not slots:
        for r_idx in range(4, 9):
            row_bg = NAVY_LIGHT if r_idx % 2 == 0 else WHITE
            for col in range(1, total_cols + 1):
                c = ws.cell(row=r_idx, column=col, value="")
                c.fill = _fill(row_bg)
                c.border = _border()
            ws.row_dimensions[r_idx].height = 22


# ── PER-PRINCIPAL ITINERARY ────────────────────────────────────────────────────

def build_principal_tab(wb, principal, p_idx, cfg):
    ws = wb.create_sheet(f"{principal['name']} — Itinerary")
    ws.freeze_panes = "A4"

    tab_color = PRINCIPAL_COLOURS[p_idx % len(PRINCIPAL_COLOURS)]
    _title_row(ws, f"Principal Itinerary — {principal['name']} | {principal['title']}", 9,
               fill_color=NAVY_DARK)

    # Meta row
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (f"Event: {cfg['event_name']}  |  {cfg['dates']}  |  {cfg['location']}  |  "
               f"POC: {principal['poc']}  |  Last Updated: [Date]")
    c.font = _font(italic=True, size=9, color=DARK_GREY)
    c.fill = _fill(NAVY_LIGHT)
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 22

    headers = ["Date", "Time (Start)", "Time (End)", "Activity",
               "Type", "Venue / Location", "Arranged by GFTN?", "Briefing Prepared?", "Notes / Requests"]
    widths  = [14, 13, 13, 36, 18, 28, 18, 18, 32]
    _header_row(ws, 3, headers, widths, fill_color=NAVY_MID)

    yn_cells = []
    for r_idx, eng in enumerate(principal.get("engagements", []), start=4):
        typ = eng.get("type", "")
        row_bg = _row_fill(r_idx, typ)
        vals = [
            eng.get("date", ""),
            eng.get("time_start", ""),
            eng.get("time_end", ""),
            eng.get("activity", ""),
            typ,
            eng.get("venue", "TBC"),
            "Yes" if eng.get("gftn_arranged") else "No",
            "Yes" if eng.get("briefing_needed") else "N/A",
            eng.get("notes", ""),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.font = _font(size=10, color=DARK_GREY, italic=(typ == "Suggested"))
            c.fill = _fill(row_bg)
            c.alignment = _align("center" if col <= 3 else "left")
            c.border = _border()
            if col in (7, 8):
                yn_cells.append(c)
        ws.row_dimensions[r_idx].height = 22

    # Blank template rows if no engagements
    if not principal.get("engagements"):
        for r_idx in range(4, 11):
            row_bg = NAVY_LIGHT if r_idx % 2 == 0 else WHITE
            for col in range(1, 10):
                c = ws.cell(row=r_idx, column=col, value="")
                c.fill = _fill(row_bg)
                c.border = _border()
            ws.row_dimensions[r_idx].height = 22

    if yn_cells:
        _add_dv(ws, '"Yes,No,N/A"', yn_cells)

    # Legend
    legend_row = max(4 + len(principal.get("engagements", [])) + 2, 13)
    ws.merge_cells(f"A{legend_row}:I{legend_row}")
    c = ws.cell(row=legend_row, column=1,
        value="Legend:  Yellow = Suggested activity (POC recommendation, not GFTN-arranged)  |  "
              "Orange = Travel / Meal  |  Rows with Arranged by GFTN? = Yes are GFTN-managed engagements")
    c.font = _font(italic=True, size=9, color=DARK_GREY)
    c.fill = _fill(NAVY_LIGHT)
    c.alignment = _align("left")
    ws.row_dimensions[legend_row].height = 20


# ── POC CHECKLIST ──────────────────────────────────────────────────────────────

def build_checklist_tab(wb, cfg):
    ws = wb.create_sheet("POC Checklist")
    _title_row(ws, "POC Checklist — All Principals", 5, fill_color=NAVY_DARK)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    _header_row(ws, 2,
        ["", "Principal / POC", "Task", "Status", "Notes"],
        [5, 24, 50, 18, 30],
        fill_color=NAVY_MID)

    # ── Comprehensive task list aligned to template (88+ tasks across 5 phases)
    sections = [
        ("Upon Assignment", [
            "Confirm the principal's role and all commitments at the event in writing",
            "Introduce yourself to the principal (warm email or WhatsApp if not already known)",
            "Get the principal's travel preferences (airline, seat, dietary requirements)",
            "Get the principal's accommodation preferences and any special requests",
            "Set up the principal's dedicated tab in the Schedule Tracker",
            "Confirm you have been added to the event POC WhatsApp group",
            "Align with Main Coordinator on scope of responsibilities",
            "Check if principal has any pre-existing contacts at the event to flag",
        ]),
        ("2–3 Weeks Before", [
            "Book flights — confirm class, airline preference, and routing",
            "Book accommodation — confirm hotel, room type, check-in/out dates",
            "Book airport-to-hotel transfer on arrival",
            "Book hotel-to-airport transfer on departure",
            "Book any inter-venue transfers during the event",
            "Schedule sidebar meetings and bilateral sessions on behalf of principal",
            "Identify at least 2 'Suggested' activities for the principal's free time (networking receptions, dinners, cultural events)",
            "Add all 'Suggested' activities to the Schedule Tracker (yellow rows)",
            "Confirm all formal GFTN-arranged engagements and lock in times and venues",
            "Begin drafting briefing documents for all confirmed engagements",
            "Check if principal needs any tech setup (slides, clicker, AV requirements)",
            "Confirm any dietary requirements with the event venue or organisers",
            "Confirm visa requirements if event is international",
        ]),
        ("1 Week Before", [
            "Send the full confirmed schedule to the principal as a clean PDF or Word doc",
            "Finalise and share all briefing documents with the principal",
            "Ensure Schedule Tracker is updated with all confirmed times, venues, and contacts",
            "Share Schedule Tracker with all POCs on SharePoint with Edit access",
            "Confirm WhatsApp group is set up and all POCs and ops contacts are added",
            "Brief your team on coverage responsibilities for each session",
            "Confirm all venues have been communicated to the principal",
            "Double-check all transport bookings and share confirmation details with principal",
            "Confirm hotel check-in time and room arrangements with accommodation",
            "Send a 'we're ready for you' message to the principal 2 days before the event",
            "Prepare a one-page day-by-day summary for the principal if schedule is complex",
            "Confirm point of contact at the venue for the day of the event",
        ]),
        ("Day of / During Event", [
            "Morning check-in with principal at start of each day (WhatsApp or in person)",
            "Ensure POC is in position at least 15 minutes before each session",
            "Accompany principal to and from all formal GFTN-arranged engagements",
            "Ensure principal has briefing documents before each engagement",
            "Flag any last-minute schedule changes immediately via the WhatsApp group",
            "Proactively suggest 'Suggested' activities to the principal during free time",
            "Manage any requests from principal in real time (meals, transport, intro requests)",
            "Update the Schedule Tracker with any changes or notes throughout the day",
            "Capture any follow-up actions or commitments made by the principal during sessions",
            "Debrief with the POC team at the end of each event day",
            "Confirm next-day arrangements with principal at end of each day",
        ]),
        ("Post-Event", [
            "Confirm principal's departure arrangements and transport to airport",
            "Ensure principal has received all necessary documents and contact details",
            "Follow up on any outstanding requests or commitments made during the event",
            "Send a thank-you note from the GFTN team to the principal",
            "Compile any follow-up actions noted during the event",
            "File all briefing documents, notes, and the final tracker in the SharePoint event folder",
            "Update the Schedule Tracker with final status of all engagements",
            "Submit any expense claims or invoices related to principal logistics",
            "Hold a team debrief — document what worked well and what to improve",
            "Archive the event WhatsApp group or download key messages for reference",
        ]),
    ]

    task_dv_cells = []
    current_row = 3
    principals = cfg["principals"]

    for section_title, tasks in sections:
        # Section header
        ws.merge_cells(f"A{current_row}:E{current_row}")
        c = ws.cell(row=current_row, column=1, value=section_title)
        c.font = _font(bold=True, size=10, color=WHITE)
        c.fill = _fill(NAVY_MID)
        c.alignment = _align("left")
        c.border = _border()
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for t_idx, task in enumerate(tasks):
            row_bg = WHITE if t_idx % 2 == 0 else PALE_GREY
            for p in principals:
                # Checkbox
                cb = ws.cell(row=current_row, column=1, value="\u2610")
                cb.font = Font(name="Calibri", size=13, color=NAVY_MID)
                cb.fill = _fill(row_bg)
                cb.alignment = _align("center", wrap=False)
                cb.border = _border()

                # Principal / POC
                p_cell = ws.cell(row=current_row, column=2,
                    value=f"{p['name']}\n(POC: {p['poc']})")
                p_cell.font = _font(size=9, color=DARK_GREY)
                p_cell.fill = _fill(row_bg)
                p_cell.alignment = _align("left")
                p_cell.border = _border()

                # Task
                task_cell = ws.cell(row=current_row, column=3, value=task)
                task_cell.font = _font(size=10)
                task_cell.fill = _fill(row_bg)
                task_cell.alignment = _align("left")
                task_cell.border = _border()

                # Status dropdown
                sc = ws.cell(row=current_row, column=4, value="To Do")
                sc.font = _font(size=10, color=DARK_GREY)
                sc.fill = _fill(row_bg)
                sc.alignment = _align("center", wrap=False)
                sc.border = _border()
                task_dv_cells.append(sc)

                # Notes
                nc = ws.cell(row=current_row, column=5, value="")
                nc.font = _font(size=10)
                nc.fill = _fill(row_bg)
                nc.alignment = _align("left")
                nc.border = _border()

                ws.row_dimensions[current_row].height = 28
                current_row += 1

        current_row += 1  # blank separator row between sections

    if task_dv_cells:
        _add_dv(ws, '"To Do,In Progress,Done,N/A"', task_dv_cells)


# ── WHATSAPP GROUP SETUP ───────────────────────────────────────────────────────

def build_whatsapp_tab(wb, cfg):
    ws = wb.create_sheet("WhatsApp Group Setup")
    _title_row(ws, "Event POC WhatsApp Group — Setup Guide", 3, fill_color=NAVY_DARK)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 36

    _header_row(ws, 2, ["Field", "Details", "Notes"], [30, 46, 36], fill_color=NAVY_MID)

    pocs_str = ", ".join(cfg.get("pocs", ["[POC Names]"]))
    rows = [
        ("Group Name",
         f"{cfg['event_name']} — POC Coordination",
         "Include event year if helpful for future reference"),
        ("Created by",
         cfg.get("coordinator", "Main Coordinator"),
         "Set up at least 1 week before the event"),
        ("Members — POCs",
         pocs_str,
         "One POC per principal — all must be added before the event"),
        ("Members — Operations",
         "Project Lead / Head of Operations",
         "For room requests, meal changes, schedule queries"),
        ("Purpose",
         "Real-time on-the-day coordination between POCs and ops",
         "Urgent issues, principal requests, last-minute changes only — use email for pre-event admin"),
        ("Admin Setting",
         "Restrict message sending to admins only (pre-event)",
         "Remove restriction on the day of the event. Prevents noise in the lead-up."),
        ("Escalation",
         f"Tag @{cfg.get('coordinator','Main Coordinator')} for cross-principal issues",
         "Tag @Operations for venue/event-level requests"),
    ]

    for r_idx, (field, detail, note) in enumerate(rows, start=3):
        row_bg = NAVY_LIGHT if r_idx % 2 == 1 else WHITE
        for col, val in enumerate([field, detail, note], start=1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.font = _font(size=10, bold=(col == 1 and bool(field)))
            c.fill = _fill(row_bg)
            c.alignment = _align("left")
            c.border = _border()
        ws.row_dimensions[r_idx].height = 22


# ── ENTRY POINT ────────────────────────────────────────────────────────────────

def generate(config_path, output_path):
    with open(config_path) as f:
        cfg = json.load(f)

    wb = Workbook()
    build_master_overview(wb, cfg)
    for p_idx, p in enumerate(cfg["principals"]):
        build_principal_tab(wb, p, p_idx, cfg)
    build_checklist_tab(wb, cfg)
    build_whatsapp_tab(wb, cfg)

    wb.save(output_path)
    print(f"Tracker saved: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate GFTN Principal Staffing Tracker")
    parser.add_argument("--config",  required=True, help="Path to event JSON config file")
    parser.add_argument("--output",  required=True, help="Output .xlsx path")
    args = parser.parse_args()
    generate(args.config, args.output)
